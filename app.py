"""
Branham Budget Tracker

created by: Anthony Branham
created on: 2/19/2021
last updated on: 3/29/2021

"""
import datetime
import tkinter
import openpyxl
import os.path
import os
import pyfiglet
from os import system, name
from calendar import monthrange
from openpyxl.chart import PieChart, Reference
from tkinter import *
from PIL import ImageTk, Image

root = Tk()
root.title(" - Branham Budget Tracker")
root.iconbitmap("images\BBT.ico")


banner = ImageTk.PhotoImage(Image.open("images\BBTheader.png"))
label1 = Label(image=banner)
label1.grid(row=0, columnspan=3)


def quit_verify():
    qw=Tk()
    frame1 = Frame(qw, highlightbackground="green", highlightcolor="green",highlightthickness=1, bd=0)
    frame1.pack()
    qw.overrideredirect(1)
    qw.geometry("200x70+650+400")
    lbl = Label(frame1, text="Are you sure you want to quit?")
    lbl.pack()
    yes_btn = Button(frame1, text="Yes", bg="red", fg="white", command=quit, width=10)
    yes_btn.pack(padx=10, pady=10 , side=LEFT)
    no_btn = Button(frame1, text="No", bg="#216d45", fg="black",command=qw.destroy, width=10)
    no_btn.pack(padx=10, pady=10, side=LEFT)
    qw.mainloop()


# Configure button functionality
cat_button = Button(root, text="Categories", font=10, width=21, height=3, bg="#216d45", fg="white")
inc_button = Button(root, text="Incomes", font=10, width=21, height=3, bg="#216d45", fg="white")
exp_button = Button(root, text="Expenses", font=10, width=21, height=3, bg="#216d45", fg="white")
placeholder1 = Label(root, text=" ")
chart_button = Button(root, text="Show chart", font=10, width=21, height=3, bg="#216d45", fg="white")
reset_button = Button(root, text="Reset data", font=10, width=21, height=3, bg="#216d45", fg="white")
quit_button = Button(root, text="Quit", font=10, width=21, height=3, bg="#216d45", fg="white", command=quit_verify)


# Configure button layout
cat_button.grid(row=1, column=0)
inc_button.grid(row=1, column=1)
exp_button.grid(row=1, column=2)
chart_button.grid(row=2, column=0)
reset_button.grid(row=2, column=1)
quit_button.grid(row=2, column=2)

root.grid
root.geometry("900x600")
root.mainloop()


# def clear_screen():
#     # Get system OS so clear works with Windows, Mac, Linux
#     if name == "nt":
#         _ = system("cls")
#     else:
#         _ = system("clear")


# def open_excel():
#     # Get system OS so clear works with Windows, Mac, Linux
#     if name == "nt":
#         _ = system("start excel.exe BudgetTracker.xlsx")
#     elif name == "posix":
#         _ = system("libreoffice --calc BudgetTracker.xlsx")
#     else:
#         _ = system("open -a 'Microsoft Excel.app' 'BudgetTracker.xlsx'")


# def print_banner():
#     # Print Banner
#     ascii_banner = pyfiglet.figlet_format("Branham Budget Tracker")
#     print(ascii_banner)
# print_banner()


# def get_name():
#     # Get user's name and generate welcome message
#     prompt_name = input("Please enter your first name: ")
#     first_name = str(prompt_name).upper()
#     print("\n\nWelcome to Branham Budget Tracker, {}!\n".format(first_name))
# get_name()


# def month_remaining():
#     # Get month and days remaining
#     today = datetime.datetime.today()
#     today_date = today.strftime("%B %d, %Y")
#     print("Today is " + today_date)
#     day_of = int(today.strftime("%d"))
#     now = datetime.datetime.now()
#     month_end = monthrange(now.year, now.month)[1]
#     days_remaining = (month_end) - (day_of)
#     print("There are {} days remaining in the month.\n".format(days_remaining))
# month_remaining()


# def create_workbook():
#     # Check for file and create if not found
#         if not os.path.exists("BudgetTracker.xlsx"):
#             wb = openpyxl.Workbook()
#             sheet = wb.active
#             sheet.title = "Budget"
#             # Create income sheet
#             ws1 = wb.create_sheet("Sheet_A")
#             ws1.title = "Income"
#             ws1.sheet_properties.tabColor = "00FF00"
#             # Create expenses sheet
#             ws2 = wb.create_sheet("Sheet_B")
#             ws2.title = "Expenses"
#             ws2.sheet_properties.tabColor = "FF0000"
#             wb.save(filename="BudgetTracker.xlsx")
#             # Create calculations sheet
#             ws3 = wb.create_sheet("Sheet_C")
#             ws3.title = "Calc"
#             ws3.sheet_properties.tabColor = "4B0082"
#             wb.save(filename="BudgetTracker.xlsx")
#         else:
#             pass
#         # Configure Budget worksheet
#         wb = openpyxl.load_workbook(filename="BudgetTracker.xlsx")
#         Budget = wb["Budget"]
#         if Budget["A1"].value is None:
#             # Create headers
#             Budget["A1"] = "CATEGORY"
#             Budget["B1"] = "PLANNED"
#             Budget["C1"] = "SPENT"
#             Budget["D1"] = "REMAINING"
#             # Create categories column
#             Budget["A2"] = "Housing"
#             Budget["A3"] = "Utilities"
#             Budget["A4"] = "Transportation"
#             Budget["A5"] = "Groceries"
#             Budget["A6"] = "Entertainment"
#             Budget["A7"] = "Debts"
#             Budget["A8"] = "Other"
#             # Create budget amount column
#             Budget["B2"] = 0
#             Budget["B3"] = 0
#             Budget["B4"] = 0
#             Budget["B5"] = 0
#             Budget["B6"] = 0
#             Budget["B7"] = 0
#             Budget["B8"] = 0
#             # Create budget amount column
#             Budget["C2"] = "=SUMIF(Expenses!C:C, A2, Expenses!A:A)"
#             Budget["C3"] = "=SUMIF(Expenses!C:C, A3, Expenses!A:A)"
#             Budget["C4"] = "=SUMIF(Expenses!C:C, A4, Expenses!A:A)"
#             Budget["C5"] = "=SUMIF(Expenses!C:C, A5, Expenses!A:A)"
#             Budget["C6"] = "=SUMIF(Expenses!C:C, A6, Expenses!A:A)"
#             Budget["C7"] = "=SUMIF(Expenses!C:C, A7, Expenses!A:A)"
#             Budget["C8"] = "=SUMIF(Expenses!C:C, A8, Expenses!A:A)"
#             # Create Remaining excel functions
#             Budget["D2"] = "=IF(B2-C2=0, \"\", B2-C2)"
#             Budget["D3"] = "=IF(B3-C3=0, \"\", B3-C3)"
#             Budget["D4"] = "=IF(B4-C4=0, \"\", B4-C4)"
#             Budget["D5"] = "=IF(B5-C5=0, \"\", B5-C5)"
#             Budget["D6"] = "=IF(B6-C6=0, \"\", B6-C6)"
#             Budget["D7"] = "=IF(B7-C7=0, \"\", B7-C7)"
#             Budget["D8"] = "=IF(B8-C8=0, \"\", B8-C8)"
#         else:
#             pass
#         # Configure Income worksheet
#         Income = wb["Income"]
#         if Income["A1"].value is None:
#         # Create headers
#             Income["A1"] = "SOURCE"
#             Income["B1"] = "AMOUNT"
#         else:
#             pass
#         # Configure Expenses worksheet
#         Expenses = wb["Expenses"]
#         if Expenses["A1"].value is None:
#         # Create headers
#             Expenses["A1"] = "AMOUNT"
#             Expenses["B1"] = "MERCHANT"
#             Expenses["C1"] = "CATEGORY"
#         # Configure Calculations worksheet
#         Calc = wb["Calc"]
#         if Calc["A1"].value is None:
#             Calc["A1"] = "Total Income"
#             Calc["B1"] = "Total Planned"
#             Calc["C1"] = "Total Spent"
#             Calc["D1"] = "Total Remaining"
#             Calc["E1"] = "$ Left to Budget"
#             Calc["A2"] = "=SUM(Income!B:B)"
#             Calc["B2"] = "=SUM(Budget!B:B)"
#             Calc["C2"] = "=SUM(Budget!C:C)"
#             Calc["D2"] = "=SUM(Budget!D:D)"
#             Calc["E2"] = "=SUM(A2-B2)"
#         else:
#             pass
#         # Adjust column widths
#         Budget.column_dimensions['A'].width = 15
#         Budget.column_dimensions['B'].width = 15
#         Budget.column_dimensions['C'].width = 15
#         Budget.column_dimensions['D'].width = 15
#         Income.column_dimensions['A'].width = 15
#         Income.column_dimensions['B'].width = 15
#         Expenses.column_dimensions['A'].width = 15
#         Expenses.column_dimensions['B'].width = 15
#         Expenses.column_dimensions['C'].width = 15
#         Calc.column_dimensions['A'].width = 15
#         Calc.column_dimensions['B'].width = 15
#         Calc.column_dimensions['C'].width = 15
#         Calc.column_dimensions['D'].width = 15
#         Calc.column_dimensions['E'].width = 15
#         wb.save("BudgetTracker.xlsx")
# create_workbook()


# def delete_workbook():
#     if os.path.exists("BudgetTracker.xlsx"):
#         os.remove("BudgetTracker.xlsx")


# def set_budget():
#     # Prompt user to select a category and set budget amount
#     wb = openpyxl.load_workbook(filename="BudgetTracker.xlsx")
#     Budget = wb["Budget"]
#     user_cat = input("1. Housing\n2. Utilities\n3. Transporation\n4. Groceries\n5. Entertainment\n6. Debts\n7. Other\nWhat category would you like to set a budget for? Enter 1-7: ").upper()
#     budget_amount = float(input("Enter budget amount: "))
#     if user_cat == "1":
#         Budget["B2"] = float(budget_amount)
#         print("Budget for HOUSING has been set to ${}.".format(budget_amount))
#     elif user_cat == "2":
#         Budget["B3"] = float(budget_amount)
#         print("Budget for UTILITIES has been set to ${}.".format(budget_amount))
#     elif user_cat == "3":
#         Budget["B4"] = float(budget_amount)
#         print("Budget for TRANSPORTATION has been set to ${}.".format(budget_amount))
#     elif user_cat == "4":
#         Budget["B5"] = float(budget_amount)
#         print("Budget for GROCERIES has been set to ${}.".format(budget_amount))
#     elif user_cat == "5":
#         Budget["B6"] = float(budget_amount)
#         print("Budget for ENTERTAINMENT has been set to ${}.".format(budget_amount))
#     elif user_cat == "6":
#         Budget["B7"] = float(budget_amount)
#         print("Budget for DEBTS has been set to ${}.".format(budget_amount))
#     elif user_cat == "7":
#         Budget["B8"] = float(budget_amount)
#         print("Budget for OTHER has been set to ${}.".format(budget_amount))
#     else:
#         print("Invalid category selection")
#         Budget
#     wb.save("BudgetTracker.xlsx")


# def mainmenu():
#     main_option = ""
#     while main_option == "":
#         """Print out of navigation menu """
#         print(30 * '-')
#         print("<<<<  MAIN MENU  >>>>")
#         print(30 * '-')
#         print("1. Category menu")
#         print("2. Income menu")
#         print("3. Expenses menu")
#         print("4. Generate breakdown")
#         print("5. Clear data/Begin new budget")
#         print("q. Quit")
#         print(30 * '-')
#         option = input("Enter an option: ")
#         clear_screen()
#         if option.lower() == 'q':
#             print("Exiting program...")
#             exit()
#         elif option == "1":
#             main_option = category_menu()
#         elif option == "2":
#             main_option = income_menu()
#         elif option == "3":
#             main_option = expenses_menu()
#         elif option == "4":
#             print("Generating breakdown...\n")
#             create_chart()
#         elif option == "5":
#             reset_verify = input("Are you sure you want to reset your budget? Y/N ").upper()
#             if reset_verify == "Y":
#                 reset_verify2 = input("**You will not be able to recover lost data after this point.**\n**Are you sure you wish to continue?** Y/N ").upper()
#                 if reset_verify2 == "Y":
#                     delete_workbook()
#                     print("All data has been reset")
#                     create_workbook()
#             elif reset_verify == "N":
#                 mainmenu()
#             else:
#                 print("Not a valid option.  Please try again.")
#                 mainmenu()
#         else:
#             print("Invalid option.  Please try again")

# # Sub menu for categories
# def category_menu():
#     cat_option = ""
#     while cat_option == "":
#         """Print out of navigation menu """
#         print(30 * '-')
#         print("<<<<  CATEGORY MENU  >>>>")
#         print(30 * '-')
#         print("1. Show categories")
#         print("2. Set budget amount")
#         print("3. Return to Main Menu")
#         print("q. Quit")
#         print(30 * '-')
#         cat_option = input("Enter an option: ")
#         clear_screen()

#         categories = []
#         values = []

#         # Open workbook
#         wb = openpyxl.load_workbook(filename="BudgetTracker.xlsx")
#         Budget = wb["Budget"]
#         # This builds our arrays
#         for row_cells in Budget.iter_rows(min_row=2, max_row=8, max_col=2):
#             for cell in row_cells:
#                 if type(cell.value) == str:
#                     categories.append(cell.value.lower())
#                 else:
#                     values.append(float(cell.value))
#         # Display current incomes
#         def display_categories():
#             print("\nCurrent categories are:\n")
#             for i in range(len(categories)):
#                 print(str(i+1) + "." + " " + categories[i].ljust(20," ") + str(values[i]))

#         if cat_option.lower() == "q":
#             print("Exiting program...")
#         # Displays categories and set budgets
#         elif cat_option == "1":
#             if len(categories) == 0:
#                 clear_screen()
#                 print("*** No categories have been added yet ***\n")
#                 category_menu()
#             else:
#                 clear_screen()
#                 display_categories()
#                 category_menu()
#         # Set budget for a category
#         elif cat_option == "2":
#             print("\nCategories are:\n")
#             set_budget()
#             category_menu()
#         # Return to main menu
#         elif cat_option == "3":
#             mainmenu()
#         else:
#             print("Invalid menu option.  Please try again")
#             category_menu()


# # Sub menu for income
# def income_menu():
#     inc_option = ""
#     while inc_option == "":
#         """Print out of navigation menu """
#         print(30 * '-')
#         print("<<<<  INCOME MENU  >>>>")
#         print(30 * '-')
#         print("1. Show Income(s)")
#         print("2. Enter new income")
#         print("3. Modify an income")
#         print("4. Delete an income")
#         print("5. return to Main menu")
#         print("q. Quit")
#         print(30 * '-')
#         inc_option = input("Enter an option: ")
#         clear_screen()

#         incomes= []
#         values = []

#         # Open workbook
#         wb = openpyxl.load_workbook(filename="BudgetTracker.xlsx")
#         Income = wb["Income"]
#         # This builds our arrays
#         for row_cells in Income.iter_rows(min_row=2, max_col=2):
#             for cell in row_cells:
#                 if type(cell.value) == str:
#                     incomes.append(cell.value.lower())
#                 else:
#                     values.append(float(cell.value))
#         # Display current incomes
#         def display_incomes():
#             print("\nCurrent Income(s) are:\n")
#             for i in range(len(incomes)):
#                 print(str(i+1) + "." + " " + incomes[i].ljust(15," ") + str(values[i]))
#         # Quit out of program
#         if inc_option.lower() == "q":
#             print("Exiting program...")
#         # Show current incomes
#         elif inc_option == "1":
#             if len(incomes) == 0:
#                 clear_screen()
#                 print("*** No incomes have been added yet ***\n")
#                 income_menu()
#             else:
#                 clear_screen()
#                 display_incomes()
#                 income_menu()
#         # Enter a new income entry
#         elif inc_option == "2":
#             display_incomes()
#             src_income = input("What is the source of this income? ")
#             if src_income.lower() in incomes:
#                 print("Source already exists")
#                 income_menu()
#             else:
#                 try:
#                     income_amount = float(input("Enter income amount: "))
#                 except ValueError:
#                     print("Not a valid amount")
#                     income_menu()
#                 Income.append([src_income, income_amount])
#                 incomes.append(src_income)
#                 values.append(income_amount)
#                 wb.save("BudgetTracker.xlsx")
#                 display_incomes()
#                 income_menu()
#         # Modify an income entry
#         elif inc_option == "3":
#             display_incomes()
#             # Prompts for new name or use existing
#             modify_income = int(input("What income would you like to edit? Enter # or 0 to return: ") or 0)
#             if modify_income == 0:
#                 income_menu()
#             new_name = str(input("New name for " + str(incomes[modify_income - 1]) + "? ") or "nothing")
#             if new_name != "nothing":
#                 incomes[modify_income - 1] = new_name
#                 Income["A" + str(modify_income + 1)] = new_name
#             # Prompts for new amount
#             new_amount = int(input("New amount? ") or 0)
#             if new_amount != 0:
#                 values[modify_income - 1] = new_amount
#                 Income["B" + str(modify_income + 1)] = float(new_amount)
#             else:
#                 print("Amount unchanged")
#             wb.save("BudgetTracker.xlsx")
#             income_menu()
#         # Delete an income entry
#         elif inc_option == "4":
#             display_incomes()
#             delete_income = int(input("What income would you like to remove? Enter # or 0 to return ") or 0)
#             if delete_income == 0:
#                 income_menu()
#             incomes.pop(delete_income - 1)
#             Income.delete_rows(delete_income + 1)
#             display_incomes()
#             wb.save("BudgetTracker.xlsx")
#             income_menu()
#         # Return to main menu
#         elif inc_option == "5":
#             mainmenu()
#         else:
#             print("Invalid menu option.  Please try again")
#             income_menu()


# # Sub menu for expenses
# def expenses_menu():
#     exp_option = ""
#     while exp_option == "":
#         """Print out of navigation menu """
#         print(30 * '-')
#         print("<<<<  EXPENSES MENU  >>>>")
#         print(30 * '-')
#         print("1. Show expense(s)")
#         print("2. Enter new expense")
#         print("3. Modify an expense")
#         print("4. Delete an expense")
#         print("5. return to Main menu")
#         print("q. Quit")
#         print(30 * '-')
#         exp_option = input("Enter an option: ")
#         clear_screen()
        
#         exp_amount = []
#         exp_merchant = []
#         exp_category = []


#         # Open workbook
#         wb = openpyxl.load_workbook(filename="BudgetTracker.xlsx")
#         Expenses = wb["Expenses"]
#         # This builds our arrays
#         for cell in Expenses["A"]:
#             exp_amount.append(cell.value)
#         for cell in Expenses["B"]:
#             exp_merchant.append(cell.value)
#         for cell in Expenses["C"]:
#             exp_category.append(cell.value)
#         # Remove Excel headers from arrays
#         exp_amount.pop(0)
#         exp_merchant.pop(0)
#         exp_category.pop(0)
#         # Display current expenses
#         def display_expenses():
#             print("\nCurrent Expense(s) are:\n")
#             for i in range(len(exp_amount)):
#                 print (str(str(i+1) + ".").ljust(5," ") +
#                 str(exp_amount[i]).ljust(12," ") +
#                 str(exp_merchant[i]).ljust(25," ") +
#                 str(exp_category[i]).ljust(20," "))
#         if exp_option.lower() == "q":
#             print("Exiting program...")
#         # Shows expenses if any entered
#         elif exp_option == "1":
#             if len(exp_amount) == 0:
#                 clear_screen()
#                 print("*** No expenses have been added yet ***\n")
#                 expenses_menu()
#             else:
#                 clear_screen()
#                 display_expenses()
#                 expenses_menu()
#         # Enter a new expense
#         elif exp_option == "2":
#             display_expenses()
#             try:
#                 expense_amount = float(input("Enter expense amount: "))
#             except ValueError:
#                 print("Not a valid amount.  Please try again")
#                 expenses_menu()
#             merch_expense = input("What is the merchant of this expense? ")
#             expense_cat = input("1. Housing\n2. Utilities\n3. Transportation\n4. Groceries\n5. Entertainment\n6. Debts\n7. Other\n\nWhat is the category for this expense? Enter 1-7:  ")
#             if expense_cat == "1":
#                 expense_cat = "Housing"
#             elif expense_cat == "2":
#                 expense_cat = "Utilities"
#             elif expense_cat == "3":
#                 expense_cat = "Transportation"
#             elif expense_cat == "4":
#                 expense_cat = "Groceries"
#             elif expense_cat == "5":
#                 expense_cat = "Entertainment"
#             elif expense_cat == "6":
#                 expense_cat = "Debts"
#             elif expense_cat == "7":
#                 expense_cat = "Other"
#             else:
#                 print("Not a valid category. Please try again")
#                 expenses_menu()
#             Expenses.append([expense_amount, merch_expense, expense_cat])
#             exp_amount.append(expense_amount)
#             exp_merchant.append(merch_expense)
#             exp_category.append(expense_cat)
#             wb.save("BudgetTracker.xlsx")
#             display_expenses()
#             expenses_menu()
#             # Modify an expense entry
#         elif exp_option == "3":
#             if len(exp_amount) == 0:
#                 clear_screen()
#                 print("*** No expenses have been added yet ***\n")
#                 expenses_menu()
#             else:
#                 display_expenses()
#             # Prompts for new amount or use existing
#             modify_expense = int(input("What expense would you like to edit? Enter # or 0 to return: ") or 0)
#             if modify_expense == 0:
#                 expenses_menu()
#             new_amount = float(input("New amount for " + (exp_merchant[modify_expense - 1]) + "? ") or 0)
#             if new_amount != 0:
#                 exp_amount[modify_expense - 1] = new_amount
#                 Expenses["A" + str(modify_expense + 1)] = float(new_amount)
#                 print("Amount changed to {}".format(new_amount))
#             else:
#                 print("Amount unchanged")
#             # Prompts for new merchant or use existing
#             new_merch = str(input("New merchant name? ") or "nothing")
#             if new_merch != "nothing":
#                 exp_merchant[modify_expense - 1] = new_merch
#                 Expenses["B" + str(modify_expense + 1)] = str(new_merch)
#                 print("Merchant has been changed")
#             else:
#                 print("Merchant unchanged")
#             # Prompts for new category or use existing
#             new_cat = str(input("New category name? ") or "nothing")
#             if new_cat != "nothing":
#                 exp_category[modify_expense - 1] = new_cat
#                 Expenses["C" + str(modify_expense + 1)] = str(new_cat)
#                 print("Category has been changed")
#             else:
#                 print("Category unchanged")
#             wb.save("BudgetTracker.xlsx")
#             expenses_menu()
#         # Delete an expense entry
#         elif exp_option == "4":
#             if len(exp_amount) == 0:
#                 clear_screen()
#                 print("*** No expenses have been added yet ***\n")
#                 expenses_menu()
#             else:
#                 display_expenses()
#             delete_expense = int(input("What income would you like to remove? Enter # or 0 to return ") or 0)
#             if delete_expense == 0:
#                 expenses_menu()
#             exp_amount.pop(delete_expense - 1)
#             Expenses.delete_rows(delete_expense + 1)
#             wb.save("BudgetTracker.xlsx")
#             display_expenses()
#             expenses_menu()
#         elif exp_option == "5":
#             mainmenu()
#         else:
#             print("Invalid menu option.  Please try again")
#             expenses_menu()


# def create_chart():
#     wb = openpyxl.load_workbook(filename="BudgetTracker.xlsx") 
#     Calc = wb["Calc"]
    
#     data = []

#     for row in Calc:
#         data.append(row)
#     # Create object of PieChart class 
#     chart = PieChart() 
#     # Create data for plotting 
#     labels = Reference(Calc, min_row = 1, max_row = 1, min_col = 2, max_col = 5) 
#     data = Reference(Calc, min_row = 2, max_row = 7, min_col = 1, max_col = 5) 
#     # Adding data to the Pie chart object
#     chart.add_data(data, from_rows=True, titles_from_data = True) 
#     # Set labels in the chart object 
#     chart.set_categories(labels) 
#     # Set the title of the chart 
#     chart.title = " Budget breakdown "
#     # Anchor chart to cell A5.
#     Calc.add_chart(chart, "A5") 
#     # save the file 
#     wb.save("BudgetTracker.xlsx")
#     open_excel()

# mainmenu()